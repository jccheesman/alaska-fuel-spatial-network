"""Build outputs/combined_friction_tables.xlsx — publication-style tables
combining the four friction-parameter tables that previously lived as separate
sheets in friction_config.xlsx, plus the per-mode cost table:

  Table 1 (sheet 'Table 1'), panels side by side to save vertical space:
    (a) Land-cover friction        (was 'LULC Friction', Table 4)
    (b) Permafrost zonal modifier  (was 'Permafrost',    Table 5)

  Table 2 (sheet 'Table 2'), stacked panels:
    (a) Ice & water constants      (was 'Ice & Water',   Table 6)
    (b) Friction-surface equations (was 'Equations',     Table 10)

  Table 3 (sheet 'Table 3'), stacked panels:
    (a) Per-mode delivery cost rates ($/gallon-mile)
    (b) Intermodal transfer fees ($/gallon)
  Values mirror friction_surface/friction_costs.py (the verified 2026-07-17
  rate set; supersedes the earlier ISER-only derivations in
  modality_cost_rates.xlsx).

Both land-cover and permafrost are three columns wide, so Table 1 places them
in adjacent column blocks separated by a spacer column, with their notes below
the panels. Styling mirrors build_input_datasets_xlsx.py (12 pt bold title,
double-border header row, thin bottom rule, wrapped top-aligned text).

Publication cleanup vs. the source sheets: internal changelog notes and
calibration-bug narration are stripped, leaving only the scientific rationale
and data sources. Numeric friction/multiplier values are shown to two
decimals. Values are otherwise unchanged from friction_config.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

# Single source of truth: pull the routing-consumed rates and fees straight
# from friction_costs.py so this workbook can never drift from the model.
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from friction_surface.friction_costs import (  # noqa: E402
    BASELINE_RATES_PER_GALLON_MILE as RATES,
    INTERMODAL_TRANSFER_FEES as FEES,
)


def _rate(mode_key: str, prec: int) -> str:
    """Per-mile rate string from friction_costs.py (prec preserves display)."""
    return f"{RATES[mode_key]:.{prec}f}"


def _fee(a: str, b: str) -> str:
    """Transfer-fee total string from friction_costs.py."""
    return "%g" % FEES[(a, b)]["total"]


OUT = Path(__file__).parent / "combined_friction_tables.xlsx"

# --- house style (matches build_input_datasets_xlsx.py) -----------------------
TITLE_FONT = Font(size=12, bold=True)
SUBTITLE_FONT = Font(size=10, italic=True)
SECTION_FONT = Font(size=11, bold=True)
HEADER_FONT = Font(size=11, bold=True)
BODY_FONT = Font(size=10)
NOTE_FONT = Font(size=8, italic=True)

DOUBLE = Side(style="double")
THIN = Side(style="thin")
HEADER_BORDER = Border(top=DOUBLE, bottom=DOUBLE)
LAST_ROW_BORDER = Border(bottom=THIN)

LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Table 1: land cover (A:C) | spacer (D) | permafrost (E:G).
T1_COL_WIDTHS = {"A": 7, "B": 20, "C": 10, "D": 3, "E": 21, "F": 14, "G": 11}
T1_NCOLS = 7
LULC_SPANS = [(1, 1), (2, 2), (3, 3)]
PERMAFROST_SPANS = [(5, 5), (6, 6), (7, 7)]

# Table 2: four columns, equations panel merges B:C for the formula.
T2_COL_WIDTHS = {"A": 22, "B": 22, "C": 20, "D": 52}
T2_NCOLS = 4
SPANS_4 = [(1, 1), (2, 2), (3, 3), (4, 4)]    # A | B | C | D      (constants)
SPANS_EQ = [(1, 1), (2, 3), (4, 4)]           # A | B:C | D        (equations)

# --- panel content (publication-cleaned) --------------------------------------
LULC_HEADERS = ("Class", "Class name", "Friction")
LULC_ROWS = [
    ("0", "water", "NoData"),
    ("1", "trees", "2.00"),
    ("2", "grass", "1.15"),
    ("3", "flooded_vegetation", "2.50"),
    ("4", "crops", "1.10"),
    ("5", "shrub_scrub", "1.60"),
    ("6", "built_area", "1.05"),
    ("7", "bare_ground", "1.10"),
    ("8", "snow_ice", "5.00"),
]
LULC_NOTE = (
    "(a) LULC friction applies only to off-network (off-road) pixels; "
    "on-network edges use ROAD_FRICTION × slope × permafrost. Values represent "
    "light-vehicle / ATV / snowmachine mobility as a continuous friction across "
    "classes. Water (class 0) is NoData in overland mode. Source: Dynamic World "
    "v1 (Brown et al. 2022)."
)

PERMAFROST_HEADERS = ("Zone (IPA convention)", "Extent fraction p", "Multiplier")
PERMAFROST_ROWS = [
    ("none / isolated", "[0.00, 0.10)", "1.00"),
    ("sporadic", "[0.10, 0.50)", "1.15"),
    ("discontinuous", "[0.50, 0.90)", "1.30"),
    ("continuous", "[0.90, 1.00]", "1.50"),
]
PERMAFROST_NOTE = (
    "(b) Applied year-round. Zone breaks (0.10, 0.50, 0.90) and the four-zone "
    "naming follow the IPA convention (Brown et al. 1997, Circum-Arctic Map of "
    "Permafrost). Per-pixel near-surface (≤1 m) permafrost probability from "
    "Pastick et al. 2015 (Remote Sens. Environ. 168, doi:10.1016/j.rse.2015.07.019). "
    "Bin labels describe the probability range at the pixel, not IPA-polygon "
    "membership."
)

ICE_HEADERS = ("Constant", "Value", "Units", "Rationale / source")
ICE_ROWS = [
    ("SEA_ICE_THRESHOLD", "0.15", "fraction [0, 1]",
     "NSIDC ice-edge definition; POLARIS limit for unreinforced barges."),
    ("RIVER_ICE_THRESHOLD", "0.15", "fraction [0, 1]",
     "Conservative operational threshold, consistent with the sea-ice value; "
     "matches Yukon / Kuskokwim practice of sailing only after breakup."),
    ("WATER_FRICTION_BARGE", "1.00", "unitless",
     "Ice-free open-water baseline; shares the reference-pixel baseline of "
     "ROAD_FRICTION. Cross-mode cost differentiation is carried separately by "
     "the per-mode $/gallon-mile rates."),
    ("ROAD_FRICTION", "1.00", "unitless",
     "Highway-grade on-road friction, combined as "
     "max(ROAD_FRICTION, f_slope) × permafrost."),
    ("ROAD_BRIDGE_FRICTION", "1.00", "unitless",
     "Highway-grade friction at bridges; slope and permafrost are skipped "
     "(engineered crossings)."),
    ("ICEROAD_TIME_PENALTY", "2.00", "× highway",
     "Loaded ice-road ≈ 25 mph vs. ≈ 50 mph highway baseline "
     "(UAF/INE 2023, Table 8.1)."),
    ("ICEROAD_BUFFER_M", "75", "m",
     "Half-pixel buffer on ice-road centerlines for 150 m grid connectivity."),
]
ICE_NOTE = None

EQ_HEADERS = ("Step", "Formula", "Where applied")
EQ_ROWS = [
    ("1a. Slope reclass",
     "f_slope = 1.00 (slope < 2°); 1.40 (2 ≤ slope < 8°); "
     "1.75 (slope ≥ 8°)",
     "Every land pixel"),
    ("1b. LULC reclass",
     "f_lulc = LULC_FRICTION[class]   (class 0 water → NoData)",
     "Every land pixel"),
    ("2a. Permafrost normalize",
     "p = p_src / 100 if max(p_src) > 1, else p_src; clip to [0, 1]",
     "Permafrost source (once)"),
    ("2b. Permafrost modifier",
     "pm = 1.00 / 1.15 / 1.30 / 1.50 for p in "
     "[0, 0.10) / [0.10, 0.50) / [0.50, 0.90) / [0.90, 1.00]",
     "Every land pixel (year-round)"),
    ("3a. Ice normalize",
     "ice = ice_src / 100 if max(ice_src) > 1, else ice_src",
     "Each ice raster (×12 per type)"),
    ("3b. Ice present",
     "sea_ice_present = sea_ice > 0.15; river_ice_present = river_ice > 0.15",
     "Per month"),
    ("4a. Static base",
     "static_base = f_slope × f_lulc; NoData on water pixels",
     "Every pixel (once); off-network base for the monthly overland stack "
     "(step 5a)"),
    ("4b. Road base",
     "road_base = max(ROAD_FRICTION, f_slope) × pm; no LULC, no water mask; "
     "NoData-free",
     "Every pixel (once); standalone static raster (road_base.tif) sampled "
     "along network road edges"),
    ("5a. Overland: base",
     "out = static_base × pm",
     "Land + valid LULC, all months"),
    ("5b. Overland: roads",
     "out = max(ROAD_FRICTION, f_slope) × pm",
     "Road ∧ land, all months"),
    ("5c. Overland: bridges",
     "out = ROAD_BRIDGE_FRICTION = 1.00",
     "Road ∧ water, all months"),
    ("5d. Overland: ice-road",
     "out = max(ROAD_FRICTION, f_slope) × 2.00 × pm",
     "Ice-road ∧ land, Jan–Mar (LULC dropped)"),
    ("6a. Barge: navigable",
     "navigable = water ∧ ¬sea_ice_present ∧ ¬river_ice_present",
     "Per month"),
    ("6b. Barge: friction",
     "out = 1.00 on navigable; NoData elsewhere",
     "Per month"),
]
EQ_NOTE = (
    "Steps are applied in sequence within the overland branch; later steps "
    "overwrite the value at their target pixels. Static base (4a) and road "
    "base (4b) are distinct products: static base carries land cover and is "
    "NoData on water, seeding the monthly overland rasters; road base drops "
    "land cover (roads through forest would otherwise read as impassable) "
    "and is NoData-free, emitted once per run as road_base.tif for "
    "network-edge sampling."
)

# Table 3: mode costs. Panel (a) four columns; panel (b) merges C:D for notes.
T3_COL_WIDTHS = {"A": 22, "B": 13, "C": 17, "D": 64}
T3_NCOLS = 4

RATE_HEADERS = ("Mode", "Rate ($/gal-mi)", "Season (model)", "Derivation / source")
RATE_ROWS = [
    ("Road", _rate("Road", 4), "Year-round",
     "Highway-tanker carriage incl. empty backhaul. Econ One transport map "
     "(Valdez–Fairbanks $0.20/gal over 364 mi; CPI 2015→26), ATRI cost "
     "build-up, and an 11-point state-contract distance regression "
     "(slope 0.00067, R² ≈ 0.75) converge at 0.0007. Range 0.00045–0.0011; "
     "Dalton corridor ≈ 0.0012–0.0015."),
    ("Barge", _rate("Barge", 4), "Ice-free (May–Oct)",
     "Blended all-in rate at Bethel-type distance. Decomposition if "
     "linehaul/river legs ever split: ocean linehaul 0.00023 (ISER market "
     "rate $0.19–0.22/gal over the 1,276-statute-mile Anchorage–Bethel "
     "route), river distribution 0.011 (DEC 2007 invoice $0.007/gal per map "
     "mile); lightering/terminal $0.8–1.2/gal fixed component currently "
     "blended into the rate."),
    ("Plane", _rate("Plane", 3), "Year-round",
     "Charter-equivalent DC-6/C-46 cost at typical 130–350 mi stage "
     "lengths. Revealed retail premiums (DCRA Jan 2026 vs. the $3.60 "
     "Fairbanks wholesale anchor), DC-6 cost build-up, ISER 2010 benchmark, "
     "and a federal charter award converge at 0.025 (range 0.015–0.042). "
     "Underprices short (<75 mi) shuttles (~0.1–0.2 real)."),
    ("Ice road", _rate("IceRoad", 3), "Winter (Jan–Mar)",
     "Single rate spanning two verified regimes (no regime split): "
     "engineered ice road with tanker ≈ 0.004 (Nuiqsut, Kuskokwim) vs. "
     "tundra snow-trail cat-train ≈ 0.02 (Atqasuk, Anaktuvuk Pass); traffic "
     "split ≈ 50/50 (PCE FY2025 gallons). Known bias ≈ 2× high for "
     "engineered edges, ≈ 2× low for trail edges."),
]
RATE_NOTE = (
    "(a) Rates in USD per gallon-mile (2026$), mirroring friction_costs.py. "
    "Each rate survived three verification rounds — blind multi-method "
    "derivation, web primary-source verification, and local-document "
    "verification (ATRI, Econ One, PCE FY2025, NSB, Noatak) — documented in "
    "Fuel Cost Blind Derivations.pdf. Gallon-mile normalization makes modes "
    "comparable despite a ~1,000× capacity spread (3,000-gal ice-road truck "
    "vs. 3M-gal linehaul barge). Rates are operational costs kept strictly "
    "separate from the environmental friction surface (baseline 1.0); "
    "seasonal availability is enforced by the friction rasters and "
    "edge-month weights, not by the rates."
)

FEE_HEADERS = ("Transfer (from ↔ to)", "Fee ($/gal)", "What it counts / note")
FEE_SPANS = [(1, 1), (2, 2), (3, 4)]
FEE_ROWS = [
    ("barge ↔ overland", _fee("barge", "overland"),
     "Road-side truck-rack out-loading only (Crowley Bethel rack $0.276 "
     "CPI-2026, shaded for margin); barge marine-header intake excluded as "
     "already in the all-in barge rate. Keys 205 edges. Was 0.40."),
    ("plane ↔ overland", _fee("plane", "overland"),
     "Road-side tarmac receiving labor + fittings only; aircraft pump and "
     "crew excluded as already in the all-in plane charter rate. Was 0.157."),
    ("overland ↔ ice road", _fee("overland", "ice_road"),
     "Road-carriage side of a continuous truck↔ice-road-truck pumped "
     "changeover; ice-road pump-in dwell excluded (in the +20% adder). "
     "Was 0.25."),
    ("barge ↔ ice road", _fee("barge", "ice_road"),
     "Thin facility-interface residual only (both incident modes all-in). "
     "Keys 8 edges but is NEVER traversed — barge May–Oct vs ice road "
     "Jan–Mar, zero month overlap. Completeness only. Was 0.40."),
]
FEE_NOTE = (
    "(b) Storage-free, rate-aware connection fees in USD per gallon "
    "(blind re-derivation, blind_derivations/06_connection_costs.md), applied "
    "once per modal handoff, direction-insensitive. Each fee bills only "
    "handling with no home in a per-mile rate — the road-carriage side, since "
    "Road is carriage-only — and excludes the all-in side (barge intake, "
    "aircraft pump/crew, ice-road +20% load/unload adder) to avoid "
    "double-counting; hub storage is excluded entirely (never a graph node). "
    "The retired storage-leg fees are documented in 05_transfer_fees.md. "
    "Total path cost ($/gal) = Σ(mode rate × miles) + the matching fee at "
    "every modal handoff."
)

# Panel (c): full citations for every source behind the rates and the fees.
SOURCE_HEADERS = ("Source", "Full citation + link / locator")
SOURCE_SPANS = [(1, 1), (2, 4)]           # A | B:D
SOURCE_ROWS = [
    ("Blind derivations",
     "Fuel Cost Blind Derivations.pdf and blind_derivations/*.md (project "
     "root): four blind per-mode rate derivations (01_road, 02_barge, "
     "03_plane, 04_ice_road) with dated verification addenda, plus the "
     "transfer-fee chapters 05_transfer_fees.md (through-storage) and "
     "06_connection_costs.md (storage-free, rate-aware). Values encoded in "
     "friction_surface/friction_costs.py."),
    ("ATRI 2025",
     "American Transportation Research Institute (2025). An Analysis of the "
     "Operational Costs of Trucking: 2025 Update. Road-rate build-up (Tables "
     "8/11: $2.260 average / $2.32 specialized per vehicle-mile); Tables 8/9 "
     "verified carriage-only (no rack / terminal / loading line) — the basis "
     "for keeping road-side handling in the connection fees."),
    ("Econ One",
     "Econ One Research (B. Pulliam), The State of Alaska's Refining Industry "
     "(Alaska DEC / RCA filing). Road rate: Valdez–Fairbanks truck transport "
     "$0.20/gal over 364 mi (CPI 2015→2026)."),
    ("State fuel contracts",
     "State of Alaska FY25/FY26 term fuel-supply contracts (heating-oil / "
     "ULSD delivered-price lists). 11-point within-vendor distance regression "
     "for the road rate (slope $0.00067/gal-mi, R² ≈ 0.75); delivered-price "
     "adders used to validate the modeled $/gal."),
    ("ISER 2010",
     "Szymoniak, N., Fay, G., Villalobos-Melendez, A., Charon, J., Smith, M. "
     "(2010). Components of Alaska Fuel Costs: An Analysis of the Market "
     "Factors and Characteristics that Influence Rural Fuel Prices. UAA "
     "Institute of Social and Economic Research, for the Alaska State "
     "Legislature, Senate Finance Committee. Barge linehaul market rate "
     "$0.19–0.22/gal; shore-side offload-labor and terminal-handling bands "
     "($0.04–0.06/gal) for the connection-fee cost build."),
    ("NOAA port distances",
     "NOAA Office of Coast Survey, Distances Between United States Ports. "
     "Anchorage–Bethel 1,109 nm = 1,276 statute mi — the barge route-distance "
     "basis."),
    ("DEC 2007 / DCCED 2005",
     "Alaska Dept. of Environmental Conservation (2007) and Dept. of "
     "Commerce, Community and Economic Development (2005) rural bulk-fuel "
     "transport studies. River-distribution invoice rate $0.007/gal per map "
     "mile."),
    ("PCE FY2025",
     "Alaska Energy Authority, Power Cost Equalization Program: Statistical "
     "Report by Community, FY2025 (final, 2026-03-01). Margin-free utility "
     "fuel prices and delivered gallons by community; plane revealed-premium "
     "anchor and ice-road ~50/50 regime split (Nuiqsut 279,232 vs Atqasuk "
     "274,877 gal)."),
    ("DCRA Jan 2026",
     "Alaska DCCED Division of Community and Regional Affairs, Alaska Fuel "
     "Price Report (Winter / January 2026) and the state ArcGIS community "
     "fuel-price service. Plane revealed delivered-price premiums vs the "
     "$3.60/gal road-served wholesale anchor."),
    ("Air tariffs & awards",
     "Everts Air Cargo tariff (verified 2026-07-15; 26% fuel surcharge); "
     "Wright Air Service and Northern Air Cargo tariffs (within 0–6%); "
     "USAspending.gov federal fuel-charter award ($11,792.50 for 1,400 gal). "
     "DC-6 / C-46 charter-equivalent plane rate."),
    ("Ice-road operations",
     "UIC Oil & Gas Support account of the Feb–Mar 2026 Anaktuvuk Pass "
     "emergency fuel haul (uicalaska.com); Northwest Arctic Borough / Native "
     "Village of Noatak Winter Fuel Haul System proposal (2012 legislative "
     "grant TPS 58673v1: $425K tracked tractor, 3,000–5,000 gal/trip, ~200 "
     "trips/season); ASTAR (Arctic Strategic Transportation and Resources) "
     "trail study; North Slope Borough FY26-27 budget; Ontario and NWT "
     "winter-road analogs incl. Tibbitt-to-Contwoyto."),
    ("Crowley Bethel tariff",
     "Crowley Fuels LLC, Schedule of Terminal Storage and Throughput Rates — "
     "Bethel, Alaska (eff. 2022-08-10): Marine Header throughput $0.1500/gal "
     "and Truck Rack $0.2500/gal (storage billed separately), CPI-adjusted to "
     "2026$. Basis for the barge↔overland connection fee (road-side rack "
     "leg)."),
    ("Port tariffs",
     "Port of Bethel Terminal Tariff #5 §206(D) inbound-petroleum wharfage "
     "($0.06/gal, 2026 column); Port of Nome Tariff No. 16.5 §05.020 "
     "bulk-liquid wharfage ($0.048/gal) and §05.035 header-crew labor "
     "($115/hr); Port of Alaska Tariff Item 260 fuel-transfer rate. Municipal "
     "wharfage / facility-interface cross-checks for the connection fees."),
    ("AK DOL Pamphlet 600",
     "Alaska Dept. of Labor & Workforce Development, Laborers' & Mechanics' "
     "Minimum Rates of Pay (Pamphlet 600, Issue 52). Fully-loaded Alaska "
     "labor rates for the operational build-up of the ice-road and tarmac "
     "connection fees (e.g. Material Handler $52.66/hr, Fueler / Oil "
     "Distributor Truck operator $84.34/hr)."),
    ("Bristol Bay Borough",
     "Bristol Bay Borough assembly records: aviation tarmac fuel-pumping fee "
     "($550 flat per service event), the prior anchor for the plane↔overland "
     "handoff (now scoped to the road-side receiving portion)."),
]
# Link / locator per source: verified URLs where one appears in the vetted
# derivation chapters, in-repo file paths for the local source PDFs, and an
# explicit "no stable public URL" where none exists (no fabricated links).
SOURCE_LINKS = {
    "Blind derivations":
        "blind_derivations/ and Fuel Cost Blind Derivations.pdf (in repo)",
    "ATRI 2025":
        "cost_derivation_resources/ATRI-Operational-Costs-of-Trucking-07-2025.pdf (in repo)",
    "Econ One":
        "cost_derivation_resources/6_barry_pulliam-_econ_one_-_the_state_of_alaskas_refining_industry.pdf (in repo)",
    "State fuel contracts":
        "https://oppm.doa.alaska.gov/media/1337/08-heating-oil.pdf ; "
        "https://oppm.doa.alaska.gov/media/1339/08-marine-diesel-fuels.pdf",
    "ISER 2010":
        "https://iseralaska.org/static/legacy_publication_links/componentsoffuel3.pdf",
    "NOAA port distances":
        "NOAA OCS, Distances Between United States Ports (no stable public URL)",
    "DEC 2007 / DCCED 2005":
        "Alaska DEC/DCCED rural fuel-transport studies (no stable public URL)",
    "PCE FY2025":
        "cost_derivation_resources/2026.03.01 FY2025 PCE Statistical Report by "
        "Community (Final).pdf (in repo)",
    "DCRA Jan 2026":
        "https://www.commerce.alaska.gov/web/Portals/4/pub/RA/Fuel_price_report/"
        "Alaska%20Fuel%20Price%20Report%20-%20January%202026.pdf ; ArcGIS: "
        "https://maps.commerce.alaska.gov/server/rest/services/Services/CDO_Utilities/MapServer",
    "Air tariffs & awards":
        "https://evertsair.com/cargo/rates ; https://wrightairservice.com/freight-prices/ ; "
        "https://www.nac.aero/rates/ ; "
        "https://www.usaspending.gov/award/CONT_AWD_140P9725F0047_1443_140P9725A0010_1443",
    "Ice-road operations":
        "https://www.uicalaska.com/2026/04/15/uicogs-emergency-fuel-haul-to-anaktuvuk-pass/ ; "
        "cost_derivation_resources/Northwest_Arctic_Borough_Noatak_Winter_Fuel_Haul_System.pdf ; "
        "cost_derivation_resources/NorthSlopeBorough-Budget-Book-FY26-27-FINAL-1.pdf (in repo)",
    "Crowley Bethel tariff":
        "Crowley Fuels Bethel terminal rate sheet (external); figures recorded "
        "in blind_derivations/05_transfer_fees.md & 06_connection_costs.md",
    "Port tariffs":
        "Port of Bethel Tariff #5 / Port of Nome Tariff 16.5 / Port of Alaska "
        "Item 260 (external); figures recorded in blind_derivations/05–06",
    "AK DOL Pamphlet 600":
        "Alaska DOL, Laborers' & Mechanics' Minimum Rates of Pay, Pamphlet 600 "
        "Issue 52 (labor.alaska.gov)",
    "Bristol Bay Borough":
        "Bristol Bay Borough assembly records (external; no stable public URL)",
}
SOURCE_NOTE = (
    "(c) Full citations, each with a link or locator (a verified public URL, "
    "an in-repo file path, or an explicit 'no stable public URL'). The rate "
    "sources support the four per-gallon-mile rates; the Crowley, port, "
    "Pamphlet-600, ISER and Bristol Bay entries additionally support the "
    "storage-free connection fees. Convergence between independent methods — "
    "not any single source — is the evidence standard; per-figure page cites "
    "live in the blind_derivations chapters."
)


def _char_width(span: tuple[int, int], widths: dict[str, float]) -> float:
    """Total Excel char-width of a (possibly merged) column span."""
    letters = [chr(ord("A") + c - 1) for c in range(span[0], span[1] + 1)]
    return sum(widths[l] for l in letters)


def _bump_row_height(ws, row: int, cells: list[tuple[str, tuple[int, int]]],
                     widths: dict[str, float]) -> None:
    """Size a row to fit the tallest wrapped cell (merged cells don't auto-fit).

    Only ever grows the row, so side-by-side panels sharing a row can each
    call this without shrinking the other's fit.
    """
    max_lines = 1
    for text, span in cells:
        if not text:
            continue
        width = _char_width(span, widths)
        # ~1.05 fudge for proportional font; +1 line per explicit break.
        est = max(1, int(len(str(text)) / max(width - 2, 1) * 1.05) + 1)
        max_lines = max(max_lines, est)
    height = 12.5 * max_lines + 3
    existing = ws.row_dimensions[row].height
    if existing is None or height > existing:
        ws.row_dimensions[row].height = height


def write_panel(ws, start_row, panel_title, headers, rows, spans, widths):
    """Write one labeled sub-panel into the column block covered by *spans*.

    Returns the row index of the last data row (the thin-rule row).
    """
    block_start = min(cs for cs, _ in spans)
    block_end = max(ce for _, ce in spans)

    # Panel title (merged across the panel's own column block).
    t = ws.cell(start_row, block_start, panel_title)
    t.font = SECTION_FONT
    t.alignment = LEFT_TOP_WRAP
    ws.merge_cells(start_row=start_row, end_row=start_row,
                   start_column=block_start, end_column=block_end)
    _bump_row_height(ws, start_row, [(panel_title, (block_start, block_end))],
                     widths)

    # Header row.
    hr = start_row + 1
    for (cs, ce), head in zip(spans, headers):
        c = ws.cell(hr, cs, head)
        c.font = HEADER_FONT
        c.alignment = LEFT_TOP_WRAP
        if cs != ce:
            ws.merge_cells(start_row=hr, end_row=hr, start_column=cs, end_column=ce)
    for col in range(block_start, block_end + 1):
        ws.cell(hr, col).border = HEADER_BORDER
    _bump_row_height(ws, hr, list(zip(headers, spans)), widths)

    # Data rows.
    for i, row_vals in enumerate(rows):
        rr = hr + 1 + i
        for (cs, ce), val in zip(spans, row_vals):
            c = ws.cell(rr, cs, val)
            c.font = BODY_FONT
            c.alignment = LEFT_TOP_WRAP
            if cs != ce:
                ws.merge_cells(start_row=rr, end_row=rr,
                               start_column=cs, end_column=ce)
        _bump_row_height(ws, rr, list(zip(row_vals, spans)), widths)

    # Thin rule under the last data row, across the panel's block.
    last = hr + len(rows)
    for col in range(block_start, block_end + 1):
        ws.cell(last, col).border = LAST_ROW_BORDER

    return last


def write_note(ws, row, note, ncols, widths):
    """Full-width table note; returns the next free row index."""
    n = ws.cell(row, 1, note)
    n.font = NOTE_FONT
    n.alignment = LEFT_TOP_WRAP
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=ncols)
    _bump_row_height(ws, row, [(note, (1, ncols))], widths)
    return row + 1


def write_caption(ws, title, subtitle, ncols, widths):
    """Table title (row 1) and subtitle (row 2), merged full width."""
    t = ws.cell(1, 1, title)
    t.font = TITLE_FONT
    t.alignment = LEFT_CENTER_WRAP
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ncols)
    sub = ws.cell(2, 1, subtitle)
    sub.font = SUBTITLE_FONT
    sub.alignment = LEFT_TOP_WRAP
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=ncols)
    _bump_row_height(ws, 2, [(subtitle, (1, ncols))], widths)


def _finish_sheet(ws, widths):
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    # Freeze the title/subtitle so the panels scroll under a fixed caption.
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = Workbook()

    # --- Table 1: land cover + permafrost, side by side -----------------------
    ws1 = wb.active
    ws1.title = "Table 1"
    write_caption(
        ws1,
        "Table 1.  Land-cover and permafrost friction factors",
        "Off-network land-cover friction and the year-round permafrost zonal "
        "modifier. Friction is unitless (baseline 1.0).",
        T1_NCOLS, T1_COL_WIDTHS,
    )
    lulc_last = write_panel(ws1, 4, "(a)  Land-cover friction (Dynamic World v1)",
                            LULC_HEADERS, LULC_ROWS, LULC_SPANS, T1_COL_WIDTHS)
    pf_last = write_panel(ws1, 4, "(b)  Permafrost zonal modifier (year-round)",
                          PERMAFROST_HEADERS, PERMAFROST_ROWS, PERMAFROST_SPANS,
                          T1_COL_WIDTHS)
    r = max(lulc_last, pf_last) + 2
    r = write_note(ws1, r, LULC_NOTE, T1_NCOLS, T1_COL_WIDTHS)
    write_note(ws1, r, PERMAFROST_NOTE, T1_NCOLS, T1_COL_WIDTHS)
    _finish_sheet(ws1, T1_COL_WIDTHS)

    # --- Table 2: ice/water constants + equations, stacked --------------------
    ws2 = wb.create_sheet("Table 2")
    write_caption(
        ws2,
        "Table 2.  Ice / water constants and friction-surface equations",
        "Ice and water friction constants, and the sequenced equations used to "
        "build the monthly overland and barge friction stacks.",
        T2_NCOLS, T2_COL_WIDTHS,
    )
    last = write_panel(ws2, 4, "(a)  Ice and water friction constants",
                       ICE_HEADERS, ICE_ROWS, SPANS_4, T2_COL_WIDTHS)
    last = write_panel(ws2, last + 2, "(b)  Friction-surface equations",
                       EQ_HEADERS, EQ_ROWS, SPANS_EQ, T2_COL_WIDTHS)
    write_note(ws2, last + 1, EQ_NOTE, T2_NCOLS, T2_COL_WIDTHS)
    _finish_sheet(ws2, T2_COL_WIDTHS)

    # --- Table 3: mode cost rates + intermodal transfer fees, stacked ---------
    ws3 = wb.create_sheet("Table 3")
    write_caption(
        ws3,
        "Table 3.  Per-mode delivery cost rates and intermodal transfer fees",
        "Operational delivery costs for the multimodal routing layer, kept "
        "separate from the environmental friction surface. Total path cost "
        "($/gal) = Σ(mode rate × miles along route) + transfer fees at each "
        "modal handoff.",
        T3_NCOLS, T3_COL_WIDTHS,
    )
    last = write_panel(ws3, 4, "(a)  Baseline delivery cost rates ($/gallon-mile)",
                       RATE_HEADERS, RATE_ROWS, SPANS_4, T3_COL_WIDTHS)
    r = write_note(ws3, last + 1, RATE_NOTE, T3_NCOLS, T3_COL_WIDTHS)
    last = write_panel(ws3, r + 1, "(b)  Intermodal transfer fees ($/gallon)",
                       FEE_HEADERS, FEE_ROWS, FEE_SPANS, T3_COL_WIDTHS)
    r = write_note(ws3, last + 1, FEE_NOTE, T3_NCOLS, T3_COL_WIDTHS)
    src_rows = [
        (tag, cit + (f"\nLink: {SOURCE_LINKS[tag]}" if SOURCE_LINKS.get(tag) else ""))
        for tag, cit in SOURCE_ROWS
    ]
    last = write_panel(ws3, r + 1, "(c)  Sources (full citations + links)",
                       SOURCE_HEADERS, src_rows, SOURCE_SPANS, T3_COL_WIDTHS)
    write_note(ws3, last + 1, SOURCE_NOTE, T3_NCOLS, T3_COL_WIDTHS)
    _finish_sheet(ws3, T3_COL_WIDTHS)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
