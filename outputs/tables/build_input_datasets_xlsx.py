"""Build outputs/input_datasets.xlsx — publication-style table of workflow input datasets.

Mirrors the styling of outputs/friction_config.xlsx: 12pt bold title row,
double-border header row, thin bottom border under the last data row,
wrapped text, top alignment, generous column widths.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

OUT = Path(__file__).parent / "input_datasets.xlsx"

TITLE_FONT = Font(size=12, bold=True)
HEADER_FONT = Font(size=11, bold=True)
BODY_FONT = Font(size=10)
SECTION_FONT = Font(size=11, bold=True)

DOUBLE = Side(style="double")
THIN = Side(style="thin")

HEADER_BORDER = Border(top=DOUBLE, bottom=DOUBLE)
LAST_ROW_BORDER = Border(bottom=THIN)

LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLUMNS = ["File", "Format", "Sources", "Description"]

SPATIAL_ROWS = [
    (
        "ak_albers_roads_merge.shp",
        "ESRI Shapefile",
        "",
        "Statewide road centerlines reprojected to NAD83 Alaska Albers (EPSG:3338); merged input for the overland routing network.",
    ),
    (
        "waterways_network_ak_albers.shp",
        "ESRI Shapefile",
        "",
        "Marine and inland waterway centerlines used as the barge-mode network.",
    ),
    (
        "airports_ak_albers.shp",
        "ESRI Shapefile",
        "",
        "Airport point locations for the plane-mode network and connector generation.",
    ),
    (
        "Ice_Roads.shp",
        "ESRI Shapefile",
        "",
        "Overland ice-road centerlines (packed-snow tundra routes) burned into the overland network during the Jan-Mar season.",
    ),
    (
        "Ports_and_Harbors.shp",
        "ESRI Shapefile",
        "",
        "Port and harbor point locations used as barge-mode network access points.",
    ),
    (
        "flight_paths_ak_albers.shp",
        "ESRI Shapefile",
        "",
        "Air corridor linework between airport pairs used by the plane-mode network.",
    ),
]

FRICTION_ROWS = [
    (
        "sea_ice_median_01-12_*_150m_EPSG3338.tif",
        "GeoTIFF (12 monthly rasters)",
        "",
        "Monthly median sea-ice concentration on the 150 m EPSG:3338 grid. Thresholded at 0.15 to gate barge-mode navigability per month.",
    ),
    (
        "river_ice_01-12.tif",
        "GeoTIFF (12 monthly rasters)",
        "",
        "Monthly river-ice probability on the 150 m EPSG:3338 grid. Thresholded at 0.50 (reserved for the future frozen-river mode; currently blocks barge).",
    ),
    (
        "dynamic_world_LULC_2022_2024_summer_mode_150m_EPSG3338.tif",
        "GeoTIFF",
        "Dynamic World v1 (Google / WRI)",
        "Summer-mode land use / land cover class (2022-2024) resampled to 150 m EPSG:3338. Drives off-network friction via LULC_FRICTION lookup.",
    ),
    (
        "FABDEM_slope_150m_EPSG3338.tif",
        "GeoTIFF",
        "FABDEM (Hawker et al. 2022)",
        "Terrain slope in degrees, derived from the FABDEM forest- and building-removed DEM, resampled to 150 m EPSG:3338. Reclassed into flat / rolling / mountain friction.",
    ),
    (
        "permafrost_probability_150m_EPSG3338.tif",
        "GeoTIFF",
        "IPA Circum-Arctic Map of Permafrost (Brown et al. 1997)",
        "Permafrost extent fraction on the 150 m EPSG:3338 grid. Binned into none / sporadic / discontinuous / continuous zones with multipliers 1.00 / 1.15 / 1.30 / 1.50, applied year-round.",
    ),
]


def write_section(ws, start_row: int, section_title: str, rows: list[tuple]) -> int:
    """Write a section heading + header + data rows; return next free row index."""
    ws.cell(row=start_row, column=1, value=section_title).font = SECTION_FONT
    ws.cell(row=start_row, column=1).alignment = LEFT_CENTER_WRAP

    header_row = start_row + 1
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = LEFT_TOP_WRAP
        cell.border = HEADER_BORDER

    for i, row_vals in enumerate(rows):
        r = header_row + 1 + i
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT_TOP_WRAP

    last_data_row = header_row + len(rows)
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=last_data_row, column=col_idx).border = LAST_ROW_BORDER

    return last_data_row + 1


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Input Datasets"

    ws.cell(row=1, column=1, value="Table 1.  Workflow input datasets").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = LEFT_CENTER_WRAP
    ws.cell(
        row=2,
        column=1,
        value="Spatial network and friction-layer input datasets used by the Alaska fuel-delivery routing pipeline.",
    ).font = BODY_FONT
    ws.cell(row=2, column=1).alignment = LEFT_CENTER_WRAP

    next_row = 4
    next_row = write_section(ws, next_row, "Section A.  Spatial network inputs", SPATIAL_ROWS)
    next_row += 1
    next_row = write_section(ws, next_row, "Section B.  Friction layer inputs", FRICTION_ROWS)

    widths = {"A": 48, "B": 28, "C": 42, "D": 70}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
