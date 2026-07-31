# -*- coding: utf-8 -*-
"""Generate modality_cost_rates.xlsx — per-modality cost-per-gallon-mile table,
derivation basis, transfer fees, verification history, and sources.

Rates mirror friction_surface/friction_costs.py (the single source of truth
consumed by routing): Road 0.0007 / Barge 0.0010 / Plane 0.025 /
IceRoad 0.010 $/gallon-mile (2026$). These are the verified rates applied
2026-07-17 after three rounds — blind multi-method derivation (2026-07-15),
web primary-source verification (2026-07-15), and local-document verification
(2026-07-17) — documented in "Fuel Cost Blind Derivations.pdf" (project root;
chapters in blind_derivations/*.md).

Supersedes the earlier ISER-2010-only build (archived in
outputs/tables/archive/); the ISER material survives inside the blind
derivations as one of several convergent methods.

Publication style matches the house workbooks (input_datasets.xlsx,
friction_config.xlsx): no fills/borders, bold "Table N." captions, a 9pt
italic-grey source line, bold headers, an Overview contents sheet, and
sequential table numbering.

Standalone: run `python outputs/tables/make_modality_cost_table.py`.
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Single source of truth: pull the routing-consumed rates and fees straight
# from friction_costs.py so this workbook can never drift from the model.
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from friction_surface.friction_costs import (  # noqa: E402
    BASELINE_RATES_PER_GALLON_MILE as RATES,
    INTERMODAL_TRANSFER_FEES as FEES,
)

OUT = "outputs/tables/modality_cost_rates.xlsx"

# --- house style ---------------------------------------------------------
TITLE = Font(bold=True, size=12)
SUB = Font(size=9, italic=True, color="595959")
HEAD = Font(bold=True, size=11)
BODY = Font(size=10)
LABEL = Font(bold=True, size=10)
LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")


def title(ws, text, source=None):
    ws["A1"] = text
    ws["A1"].font = TITLE
    ws["A1"].alignment = LEFT
    if source:
        ws["A2"] = source
        ws["A2"].font = SUB
        ws["A2"].alignment = LEFT


def header_row(ws, row, cols):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = HEAD
        c.alignment = LEFT


def data_rows(ws, start, rows, numfmt=None):
    for r, row in enumerate(rows, start=start):
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY
            c.alignment = LEFT
            if numfmt and i in numfmt and isinstance(val, (int, float)):
                c.number_format = numfmt[i]
    return start + len(rows)


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def footnotes(ws, row, ncols, lines):
    for i, line in enumerate(lines):
        cell = ws.cell(row=row + i, column=1, value=line)
        cell.font = SUB
        cell.alignment = LEFT
        if ncols > 1:
            ws.merge_cells(start_row=row + i, start_column=1,
                           end_row=row + i, end_column=ncols)


wb = Workbook()

# ===========================================================================
# Table 1 — Overview / contents
# ===========================================================================
ws = wb.active
ws.title = "Overview"
title(ws, "Table 1.  Modality cost rates — overview",
      "Per-modality fuel delivery cost rates for the multimodal fuel-network "
      "routing layer, mirroring friction_costs.py (verified rate set, applied "
      "2026-07-17). Cost rates are the operational $/gallon premium and are "
      "kept separate from the environmental friction surface (baseline 1.0).")
header_row(ws, 4, ["Sheet", "Contents"])
end = data_rows(ws, 5, [
    ["Cost per gal-mile", "Verified per-modality rate ($/gallon-mile), range, "
     "confidence, model season, and basis of convergence"],
    ["Derivation", "The independent methods behind each rate and their "
     "verified key figures"],
    ["Transfer fees", "Intermodal transfer fees ($/gallon) charged at modal "
     "handoffs"],
    ["Veh-mile reference", "Raw vehicle-mile carrier rates and capacities "
     "(reference only, not consumed by routing)"],
    ["Verification", "The three verification rounds and the prior-model vs. "
     "verified comparison"],
    ["Sources", "Full citations"],
])
footnotes(ws, end + 1, 2, [
    "Rates are unit costs and apply per gallon delivered; no per-community "
    "delivered-volume data is required to use them.",
    "Sea/river-ice seasonality is captured in the friction rasters and "
    "edge-month weights, not in these cost rates (avoids double-counting).",
    "Every derivation found a two-part cost structure (per-mile rate + fixed "
    "per-delivery cost); the fixed components are carried as intermodal "
    "transfer fees, so a flat $/gal-mi rate is calibrated to each mode's "
    "typical route length.",
])
widths(ws, [24, 96])

# ===========================================================================
# Table 2 — Cost per gallon-mile by modality (verified)
# ===========================================================================
ws = wb.create_sheet("Cost per gal-mile")
title(ws, "Table 2.  Cost per gallon-mile by modality (verified 2026$)",
      "Point estimates from three or more genuinely independent methods per "
      "mode; convergence between methods, not any single source, is the "
      "evidence standard. Fuel Cost Blind Derivations.pdf, executive summary.")
header_row(ws, 4, ["Modality", "$ / gallon-mile", "Range", "Confidence",
                   "Season (model)", "Basis of convergence"])
end = data_rows(ws, 5, [
    ["Road (highway tanker)", RATES["Road"], "0.00045-0.0011", "Moderate",
     "Year-round",
     "Econ One transport map (Valdez-Fairbanks $0.20/gal truck); ATRI 2025 "
     "build-up (Tables 8/11: $2.260 avg / $2.32 specialized per vehicle-mile); "
     "11-point within-vendor state-contract distance regression (slope "
     "0.00067, R^2 ~ 0.75); USDOT BCA cross-check. Dalton corridor runs "
     "~0.0012-0.0015."],
    ["Barge (blended all-in)", RATES["Barge"], "see decomposition", "Med-high",
     "Ice-free (May-Oct)",
     "Blended all-in rate at Bethel-type distance for the current one-rate "
     "design. Decomposition if linehaul/river edges ever split: ocean "
     "linehaul 0.00023 (range 0.00017-0.00039) and river distribution 0.011 "
     "(range 0.006-0.018); lightering/terminal ~$0.8-1.2/gal is a fixed "
     "transfer component currently smeared into the blend."],
    ["Plane (bulk fuel by air)", RATES["Plane"], "0.015-0.042", "Med-high",
     "Year-round",
     "DCRA Jan 2026 revealed premiums vs. the empirical $3.60/gal road-served "
     "wholesale anchor (Circle/Central, PCE FY2025); DC-6 cost build-up; ISER "
     "2010 benchmark (p. 13); exact federal charter award ($11,792.50 / 1,400 "
     "gal); PCE two-part regression (fixed $1.31/gal + $0.0082/gal-mi); "
     "Wright Air and NAC tariffs corroborate Everts within 0-6%. Underprices "
     "<75-mi shuttles (~0.1-0.2 real)."],
    ["Ice road / winter trail", RATES["IceRoad"], "0.004-0.025", "Moderate",
     "Winter (Jan-Mar)",
     "Build-up calibrated to the verified 2026 Anaktuvuk haul; Noatak "
     "haul-system proposal corroborates the physics (3,000-5,000 gal/trip, "
     "10-12 mph, $425K tracked tractor); within-borough DCRA differentials; "
     "Ontario/NWT analogs and Tibbitt-Contwoyto economics. Single rate spans "
     "two regimes - engineered ice road ~0.004 (Nuiqsut, Kuskokwim) vs. "
     "tundra snow trail ~0.02 (Atqasuk, Anaktuvuk Pass) - with the ~50/50 "
     "traffic split data-backed by PCE FY2025 gallons."],
], numfmt={2: "0.00000"})
footnotes(ws, end + 1, 6, [
    "Mode ordering is economically coherent: ocean barge < road < ice road < "
    "air, each step roughly 3-40x the previous, reproducing observed "
    "behavior (barge what you can, truck what the road reaches, build winter "
    "roads to displace airlift, fly only what you must).",
    "The ice-road single rate is a confirmed midpoint, not a coin-flip, but "
    "still misprices each regime by ~2x in opposite directions (high for "
    "engineered edges, low for trail edges). Last unverified input: the "
    "~$900/hr convoy rate.",
    "Season column shows the months the mode is traversable in the model; "
    "gating is enforced by the friction rasters / edge-month weights.",
])
widths(ws, [26, 14, 16, 12, 16, 80])

# ===========================================================================
# Table 3 — Derivation methods and verified key figures
# ===========================================================================
ws = wb.create_sheet("Derivation")
title(ws, "Table 3.  Independent methods behind each rate",
      "Each mode was derived by a separate agent working blind (no access to "
      "the model's prior values or to the other agents), using at least three "
      "independent methods. Key figures below were verified against the "
      "primary documents; page cites in the chapter verification addenda.")
header_row(ws, 4, ["Modality", "Method", "Verified key figures"])

groups = [
    ("Road", [
        ("Published transport-cost filing",
         "Econ One refining-industry study: Valdez-Fairbanks truck transport "
         "$0.20/gal over 364 mi (read off the source deck), CPI 2015->26."),
        ("Engineering cost build-up",
         "ATRI 2025 Operational Costs of Trucking, Tables 8/11: $2.260 "
         "average / $2.32 specialized per vehicle-mile; 9,000-13,000-gal "
         "tanker payloads; carriage cost includes empty backhaul."),
        ("Revealed-price regression",
         "11-point within-vendor state heating-oil-contract distance "
         "regression: slope $0.00067/gal-mi, R^2 ~ 0.75."),
    ]),
    ("Barge", [
        ("Tariff / market-rate build-up",
         "ISER 2010 (p. 17): linehaul market rate $0.19-0.22/gal; NOAA "
         "Distances Between United States Ports: Anchorage-Bethel 1,109 nm = "
         "1,276 statute mi -> ocean linehaul 0.00023 $/gal-mi."),
        ("Invoice-based river rate",
         "DEC 2007 (p. 21): $0.007/gal per map mile for river distribution "
         "-> 0.011 $/gal-mi (2026$); ISER small-barge cost $0.60/gal typical "
         "(p. 20) corroborates."),
        ("Revealed-price bound",
         "PCE FY2025 margin-free utility fuel prices bound freight from "
         "above; Sleetmute $4.49/gal vs. Bethel $4.45 shows ~310 river miles "
         "adds essentially zero realized freight (procurement pooling)."),
    ]),
    ("Plane", [
        ("Revealed delivered-price premiums",
         "DCRA Jan 2026 community prices vs. the empirical $3.60/gal "
         "road-served wholesale anchor (Circle/Central, PCE FY2025); PCE "
         "two-part regression: fixed $1.31/gal + $0.0082/gal-mi at the "
         "efficient full-load end."),
        ("Engineering cost build-up",
         "DC-6 operating-cost build-up at typical 130-350 mi stage lengths; "
         "~4,000-gal (28,000-lb) diesel payload."),
        ("Published tariffs and awards",
         "Everts tariff ($0.80-2.12/lb + 26% fuel surcharge, verified "
         "2026-07-15); Wright Air and NAC within 0-6%; exact federal charter "
         "award $11,792.50 for 1,400 gal; ISER 2010 benchmark (p. 13)."),
    ]),
    ("Ice road", [
        ("Engineering cost build-up",
         "Convoy build-up calibrated to the verified 2026 Anaktuvuk haul; "
         "last load-bearing assumption is the ~$900/hr convoy rate (a real "
         "NSB/UIC contract figure would retire it)."),
        ("Operational corroboration",
         "Noatak Winter Fuel Haul System proposal: 20-30 mi route, "
         "3,000-5,000 gal/trip, 10-12 mph, ~200 trips/season, $425K tracked "
         "tractor - full operation specified but no operating rate "
         "published; ASTAR (267 pp.) and the UIC haul account also publish "
         "no fuel-haul dollars."),
        ("Analogs and differentials",
         "Within-borough DCRA price differentials; Ontario and NWT winter- "
         "road analogs; Tibbitt-Contwoyto economics. Regime traffic split "
         "data-backed ~50/50 (PCE FY2025 gallons: Nuiqsut 279,232 vs. "
         "Atqasuk 274,877)."),
    ]),
]
r = 5
for modality, items in groups:
    for i, (method, figures) in enumerate(items):
        ws.cell(row=r, column=1, value=modality if i == 0 else "").font = \
            LABEL if i == 0 else BODY
        ws.cell(row=r, column=1).alignment = LEFT
        for col, val in ((2, method), (3, figures)):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY
            c.alignment = LEFT
        r += 1
footnotes(ws, r + 1, 3, [
    "Cross-cutting finding: retail price differentials overstate freight - "
    "the majority of a remote community's premium is storage, working "
    "capital, and retail margin (e.g., ~85-90% of Bethel's $2.78/gal net "
    "premium is non-freight). Differentials were used only where non-freight "
    "components could be netted out or bounded.",
    "All rates are calibrated to pre-shock (Jan-early-2026) price "
    "relationships in real 2026 dollars; the June 2026 fuel price spike is a "
    "commodity effect, not a freight effect (~+$0.02/gal on a barge voyage).",
])
widths(ws, [14, 34, 84])

# ===========================================================================
# Table 4 — Intermodal transfer fees
# ===========================================================================
ws = wb.create_sheet("Transfer fees")
title(ws, "Table 4.  Intermodal connection (transfer) fees ($/gallon)",
      "Storage-free, rate-aware values (blind re-derivation 2026-07-28; "
      "blind_derivations/06_connection_costs.md). Each fee prices the atomic "
      "modal-boundary crossing and bills ONLY handling with no home in a "
      "per-mile rate: the road-carriage side (Road rate is carriage-only), "
      "excluding the all-in side (barge intake, aircraft pump/crew, ice-road "
      "+20% load/unload adder) to prevent double-counting. Applied once per "
      "handoff, direction-insensitive. Total path cost ($/gal) = sum(mode "
      "rate x miles) + the matching fee at every handoff. Mirrors "
      "INTERMODAL_TRANSFER_FEES in friction_costs.py.")
header_row(ws, 4, ["Transfer (from <-> to)", "Fee ($/gal)", "What it counts "
                   "/ note (prior through-storage value)"])
end = data_rows(ws, 5, [
    ["barge <-> overland", FEES[("barge", "overland")]["total"],
     "Road-side truck-rack out-loading only (Crowley Bethel rack $0.276 "
     "CPI-2026, shaded for sole-provider margin). Barge marine-header intake "
     "excluded (already in the all-in barge rate). Range 0.15-0.30, high "
     "confidence. Keys 205 graph Transfer edges. Was 0.40 (0.15 was a "
     "double-count)."],
    ["plane <-> overland", FEES[("plane", "overland")]["total"],
     "Road/ground-side tarmac receiving labor + fittings only. Aircraft "
     "onboard pump and flight/ground crew excluded (already in the all-in "
     "plane charter rate). Range 0.015-0.045, medium confidence. Latent "
     "boundary (plane<->overland handoff at air-served hubs). Was 0.157."],
    ["overland <-> ice road", FEES[("overland", "ice_road")]["total"],
     "Road-carriage side of a continuous truck<->ice-road-truck pumped "
     "changeover. Ice-road-side pump-in dwell excluded (in the 0.010 rate's "
     "+20% adder). Range 0.014-0.032, medium confidence. Latent Dalton- "
     "corridor boundary. Was 0.25 (a storage/tank-rack analog)."],
    ["barge <-> ice road", FEES[("barge", "ice_road")]["total"],
     "Thin vessel<->land-tanker facility-interface residual only; both "
     "incident modes are all-in, so near zero. Range 0.00-0.02, low "
     "confidence. Keys 8 graph edges but is NEVER traversed (barge May-Oct "
     "vs ice road Jan-Mar; zero month overlap) - completeness only. Was 0.40 "
     "(double-counted both all-in legs)."],
], numfmt={2: "0.000"})
footnotes(ws, end + 1, 3, [
    "Construction rule: bill only handling with no home in a per-mile rate. "
    "Because Road is carriage-only, that is the road-carriage side of every "
    "crossing; the all-in side is excluded (double-count guard) and hub "
    "storage is excluded entirely (never a graph node).",
    "The storage pseudo-mode legs (barge->storage, storage->overland/"
    "ice_road, plane->drums) are retired from the fee table and documented "
    "in blind_derivations/05_transfer_fees.md; they return only if Phase 6 "
    "makes hub storage an explicit node.",
    "Removing the storage leg does not change the four per-gallon-mile rates "
    "or their includes-basis (coupling proof: 06_connection_costs.md, S6.6).",
])
widths(ws, [22, 12, 88])

# ===========================================================================
# Table 5 — Raw vehicle-mile reference rates
# ===========================================================================
ws = wb.create_sheet("Veh-mile reference")
title(ws, "Table 5.  Raw vehicle-mile carrier rates (reference only)",
      "Source tariff data underlying the gallon-mile normalization; not "
      "consumed by routing. Mirrors VEHICLE_MILE_RATES_REFERENCE in "
      "friction_costs.py. Gallon-mile normalization makes modes comparable "
      "despite a ~1,000x capacity spread.")
header_row(ws, 4, ["Modality", "Vehicle-mile rate (low / point / high)",
                   "Capacity (gal)", "Note"])
end = data_rows(ws, 5, [
    ["Road", "$3.50 / $4.75 / $6.00 per veh-mi", "9,000",
     "AK heavy-transport range (heavyequipmenttransport.com); tanker "
     "payloads 9,000-13,000 gal (Dalton semis / A-trains)."],
    ["Barge", "$0.29/gal linehaul (point)", "3,000,000",
     "Anchorage->Bethel over 1,276 statute mi (NOAA 1,109 nm); ISER 2010 "
     "$0.19-0.22/gal x 1.51 CPI."],
    ["Plane", "$0.80-2.12/lb + 26% fuel surcharge", "4,000",
     "Everts tariff verified 2026-07-15; Wright Air within 0-6%; DC-6 "
     "diesel payload ~4,000 gal (28,000 lb) at 7.05 lb/gal; bulk fuel "
     "actually flies charter-only."],
    ["Ice road", "$10 / $20 / $30 per veh-mi", "3,000",
     "Engineering estimate - no published source exists; Noatak proposal "
     "confirms 3,000-5,000 gal loads at 10-12 mph but publishes no rate."],
])
widths(ws, [12, 34, 14, 66])

# ===========================================================================
# Table 6 — Verification history and prior-model comparison
# ===========================================================================
ws = wb.create_sheet("Verification")
title(ws, "Table 6.  Verification history",
      "Three rounds; convergence between independent methods is the evidence "
      "standard. All four point estimates survived all rounds.")
header_row(ws, 4, ["Round", "Date", "What was done", "Outcome"])
end = data_rows(ws, 5, [
    ["1. Blind derivation", "2026-07-15",
     "Four agents, one per mode, working blind (no access to model values, "
     "prior derivations, or each other); >= 3 independent methods each, with "
     "full derivation logs and arithmetic.",
     "Point estimates: Road 0.0007, Barge 0.00023 linehaul / 0.011 river, "
     "Plane 0.025, Ice road 0.010."],
    ["2. Web primary-source verification", "2026-07-15",
     "Fetched every flagged primary document (DCRA Winter 2026 report + "
     "ArcGIS service, state heating-oil contract, DEC/Econ One filing, NOAA "
     "port distances, ISER 2010, DCCED 2005, DEC 2007, Everts/Wright/NAC "
     "tariffs, USAspending charter award, UIC haul account, ASTAR study); "
     "verified or corrected every load-bearing number in place.",
     "All four estimates survived; barge per-mile figures shifted ~10% on "
     "the NOAA-verified route distance; snippet-era errors corrected and "
     "logged in dated chapter addenda."],
    ["3. Local-document verification", "2026-07-17",
     "Read manually downloaded documents the web round could not reach: "
     "ATRI 2025, Econ One study, AEA PCE FY2025 Statistical Report, state "
     "marine-diesel contract, NSB FY26-27 budget, OMB CWAT sheet, Noatak "
     "Winter Fuel Haul System proposal.",
     "All four estimates held. Plane wholesale anchor replaced with the "
     "empirical $3.60/gal (road-served PCE communities); ice-road 50/50 "
     "regime split confirmed by PCE gallons; Sleetmute $4.49 vs. Bethel "
     "$4.45 showed river miles add ~zero realized freight."],
    ["Applied to model", "2026-07-17",
     "Verified rates written to BASELINE_RATES_PER_GALLON_MILE in "
     "friction_costs.py (single ice-road rate, no regime split, per user "
     "decision).",
     "Road 0.0007 / Barge 0.0010 / Plane 0.025 / IceRoad 0.010."],
])

end += 1
ws.cell(row=end, column=1,
        value="Prior model values vs. verified estimates").font = LABEL
end += 1
header_row(ws, end, ["Modality", "Prior (pre-2026-07-17)", "Verified",
                     "Assessment"])
end = data_rows(ws, end + 1, [
    ["Road", "0.000528", "0.0007",
     "Prior was at the pure-carriage floor of the derived range; ~25% low "
     "vs. point."],
    ["Barge", "0.000556", "0.0010",
     "Prior sat between linehaul-only and all-in; NOAA-verified route is "
     "1,276 statute mi, not 1,800."],
    ["Plane", "0.033", "0.025",
     "Prior ~30% above the verified point; inside the range (upper half)."],
    ["IceRoad", "0.00667", "0.010",
     "Prior ~35% below the verified point; inside the range (lower half)."],
])
footnotes(ws, end + 1, 4, [
    "The prior values traced to the ISER-2010-only derivation (archived in "
    "outputs/tables/archive/); the ISER material survives inside the blind "
    "derivations as one convergent method among several.",
])
widths(ws, [26, 22, 50, 50])

# ===========================================================================
# Table 7 — Sources
# ===========================================================================
ws = wb.create_sheet("Sources")
title(ws, "Table 7.  Sources")
header_row(ws, 4, ["Tag", "Full citation"])
data_rows(ws, 5, [
    ["Blind Derivations", "Fuel Cost Blind Derivations.pdf (project root; "
     "chapters and derivation logs in blind_derivations/*.md). Four blind "
     "per-mode derivations with dated verification addenda; rates encoded in "
     "friction_surface/friction_costs.py."],
    ["ATRI 2025", "American Transportation Research Institute (2025). An "
     "Analysis of the Operational Costs of Trucking. Tables 8 and 11."],
    ["Econ One", "Econ One Research, Alaska refining-industry study (DEC "
     "filing): Valdez-Fairbanks truck transport component."],
    ["DCRA 2026", "Alaska DCCED Division of Community and Regional Affairs, "
     "Alaska Fuel Price Report (Winter/Jan 2026) and the state ArcGIS "
     "community fuel-price service."],
    ["PCE FY2025", "Alaska Energy Authority, Power Cost Equalization "
     "Program: Statistical Report FY2025 (margin-free utility fuel prices "
     "and gallons by community)."],
    ["NOAA Distances", "NOAA, Distances Between United States Ports "
     "(Anchorage-Bethel 1,109 nm)."],
    ["ISER 2010", "Szymoniak, N., Fay, G., Villalobos-Melendez, A., Charon, "
     "J., Smith, M. (2010). Components of Alaska Fuel Costs: An Analysis of "
     "the Market Factors and Characteristics that Influence Rural Fuel "
     "Prices. UAA ISER, for the Alaska State Legislature, Senate Finance "
     "Committee."],
    ["DEC 2007 / DCCED 2005", "Alaska DEC (2007) and DCCED (2005) rural "
     "fuel-transport studies (river-distribution invoice rate)."],
    ["Air tariffs", "Everts Air Cargo tariff (verified 2026-07-15, 26% fuel "
     "surcharge); Wright Air Service and Northern Air Cargo tariffs; "
     "USAspending.gov federal fuel-charter award ($11,792.50 / 1,400 gal)."],
    ["Ice-road operations", "Northwest Arctic Borough, Noatak Winter Fuel "
     "Haul System proposal; ASTAR (Arctic Strategic Transportation and "
     "Resources) trail study; UIC winter haul account; NSB FY26-27 budget; "
     "Ontario/NWT winter-road analogs incl. Tibbitt-Contwoyto."],
    ["Transfer fees", "Crowley Bethel terminal tariff (marine header, truck "
     "rack); Bristol Bay Borough assembly records (tarmac pumping fee)."],
])
widths(ws, [22, 100])

wb.save(OUT)
print("wrote", OUT)
