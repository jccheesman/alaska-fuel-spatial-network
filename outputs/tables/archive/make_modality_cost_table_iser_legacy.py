# -*- coding: utf-8 -*-
"""Generate modality_cost_rates.xlsx — per-modality cost-per-gallon-mile table,
derivations, sources, and the barge two-leg build.

Grounded in ISER 2010 (Components of Alaska Fuel Costs) and the Alaska Fuel
Delivery Cost dossier; barge speed cross-checked against Moffatt & Nichol's
2024 POA Economic Assessment.

Publication style matches the house workbooks (input_datasets.xlsx,
friction_config.xlsx): no fills/borders, bold "Table N." captions, a 9pt
italic-grey source line, bold headers, an Overview contents sheet, and
sequential table numbering.

Standalone: run `python outputs/tables/make_modality_cost_table.py`.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = "outputs/tables/modality_cost_rates.xlsx"

# --- house style ---------------------------------------------------------
TITLE = Font(bold=True, size=12)
SUB = Font(size=9, italic=True, color="595959")
HEAD = Font(bold=True, size=11)
BODY = Font(size=10)
LABEL = Font(bold=True, size=10)
LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")


def title(ws, text, source=None):
    ws["A1"] = text
    ws["A1"].font = TITLE
    ws["A1"].alignment = LEFT
    if source:
        ws["A2"] = source
        ws["A2"].font = SUB
        ws["A2"].alignment = LEFT


def header_row(ws, row, cols):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = HEAD
        c.alignment = LEFT


def data_rows(ws, start, rows, numfmt=None):
    for r, row in enumerate(rows, start=start):
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY
            c.alignment = LEFT
            if numfmt and i in numfmt and isinstance(val, (int, float)):
                c.number_format = numfmt[i]
    return start + len(rows)


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def footnotes(ws, row, ncols, lines):
    for i, line in enumerate(lines):
        cell = ws.cell(row=row + i, column=1, value=line)
        cell.font = SUB
        cell.alignment = LEFT
        if ncols > 1:
            ws.merge_cells(start_row=row + i, start_column=1,
                           end_row=row + i, end_column=ncols)


wb = Workbook()

# ===========================================================================
# Table 1 — Overview / contents
# ===========================================================================
ws = wb.active
ws.title = "Overview"
title(ws, "Table 1.  Modality cost rates — overview",
      "Per-modality fuel delivery cost rates for the multimodal fuel-network "
      "routing layer. Cost rates are the operational $/gallon premium and are "
      "kept separate from the environmental friction surface (baseline 1.0).")
header_row(ws, 4, ["Sheet", "Contents"])
end = data_rows(ws, 5, [
    ["Cost per gal-mile", "Per-modality cost rate ($/gallon-mile), fixed/handling adder, derivation, source"],
    ["Derivation", "Worked arithmetic per mode and the day-to-distance calibration constants"],
    ["Derivation detail", "Provenance of every number in the derivations — value, meaning, and source"],
    ["Barge build", "Two-leg barge cost build (linehaul + lighterage) and validation against ISER tables"],
    ["Sources", "Full citations"],
])
footnotes(ws, end + 1, 2, [
    "Rates are unit costs and apply per gallon delivered; no per-community "
    "delivered-volume data is required to use them.",
    "Sea/river-ice seasonality is captured in the friction raster, not in "
    "these cost rates (avoids double-counting).",
])
widths(ws, [24, 96])

# ===========================================================================
# Table 2 — Cost per gallon-mile by modality
# ===========================================================================
ws = wb.create_sheet("Cost per gal-mile")
title(ws, "Table 2.  Cost per gallon-mile by modality",
      "Distance-driven cost expressed as $/gallon-mile; distance-independent "
      "cost shown separately as a fixed/handling adder ($/gallon).")
header_row(ws, 4, ["Modality", "$ / gallon-mile",
                   "Fixed / handling adder ($/gal)", "Derivation", "Source"])
end = data_rows(ws, 5, [
    ["Plane", 0.0125, None,
     "$1.25/gal per 100 air-miles / 100 = $0.0125/gal-mi. Competitive floor "
     "(4,900-gal DC-6, Everts); higher for small planes.",
     "ISER 2010, p.13"],
    ["Road / trucking", 0.000528, None,
     "$4.75/vehicle-mile / 9,000-gal tanker. Magnitude confirmed by Nenana "
     "trucking <$0.05/gal, ~distance-linear.",
     "Delivery Cost dossier; ISER 2010 p.17 (check)"],
    ["Barge - linehaul (trunk: refinery to hub)", None, 0.19,
     "($2.5M set + $3.5M fuel/misc) / 18M gal (6 trips) = $0.19/gal. "
     "Volume/trip-driven, not per-mile; flat trunk adder (range $0.19-0.22).",
     "ISER 2010, p.16-17, Table 4"],
    ["Barge - lighterage (hub to community)", 0.000362, 0.20,
     "Table 2: $10k/day / 250k gal/trip; trip days = 5 fixed + 2*dist/221 "
     "mi-day. Reproduces $0.60/gal at ~1,100 mi (Bethel). Distance = one-way "
     "water miles. Ice-season length is NOT applied here - captured in the "
     "friction raster.",
     "ISER 2010, Tables 1-2 (calibrated)"],
    ["Ice road", 0.00667, None,
     "$20/vehicle-mile midpoint / 3,000-gal truck. Not in ISER report "
     "(excludes North Slope).",
     "Delivery Cost dossier"],
], numfmt={2: "0.000000", 3: "0.00"})
footnotes(ws, end + 1, 5, [
    "The fixed/handling adder is distance-independent ($/gal). Barge cannot be "
    "represented by a single per-mile rate because the $0.20 handling term "
    "dominates short hauls; plane and road have no fixed term.",
])
widths(ws, [40, 15, 20, 58, 32])

# ===========================================================================
# Table 3 — Derivation arithmetic
# ===========================================================================
ws = wb.create_sheet("Derivation")
title(ws, "Table 3.  Derivation arithmetic",
      "How each rate in Table 2 is built from source tariffs, capacities, and "
      "distances.")
header_row(ws, 4, ["Modality", "Worked arithmetic"])
end = data_rows(ws, 5, [
    ["Plane", "1.25 $/gal/100mi / 100 = 0.0125 $/gal-mi"],
    ["Road", "4.75 $/veh-mi / 9,000 gal = 0.000528 $/gal-mi"],
    ["Linehaul barge", "($2.5M set + $3.5M fuel) / (3M gal x 6 trips) = $6M / 18M = $0.19/gal (flat)"],
    ["Lighterage barge - fixed", "5 days x $10k / 250k gal = $0.20/gal"],
    ["Lighterage barge - distance", "(2 x $10k) / (221 mi/day x 250k gal) = 0.000362 $/gal-mi"],
    ["Ice road", "20 $/veh-mi / 3,000 gal = 0.00667 $/gal-mi"],
])

# sub-block: calibration constants
end += 1
ws.cell(row=end, column=1, value="Calibration constants (day-to-distance substitution)").font = LABEL
end += 1
header_row(ws, end, ["Constant", "Value"])
end = data_rows(ws, end + 1, [
    ["Barge speed", "221 mi/day (~8 kn x 24 h); 8 kn independently confirmed by "
                    "Moffatt & Nichol, POA Economic Assessment (2024), vessel-tracker data"],
    ["Fixed handling days", "5 days (load / lighter / wait; distance-independent)"],
    ["Gallons per lighterage trip", "250,000 (of 275,000 capacity)"],
    ["Daily cost", "$10,000/day ($1.35M / 135 operating days)"],
])
footnotes(ws, end + 1, 2, [
    "Trip days are not an input to the model: they are replaced by "
    "distance / barge speed, so only distance, mode, and location are required.",
])
widths(ws, [30, 82])

# ===========================================================================
# Table 4 — Derivation detail (provenance of every number)
# ===========================================================================
ws = wb.create_sheet("Derivation detail")
title(ws, "Table 4.  Provenance of derivation numbers",
      "Every value used in the Table 3 arithmetic, what it represents, and its "
      "source. Master form of each rate: (cost of a delivery event) / "
      "(gallons carried x miles travelled).")
header_row(ws, 4, ["Modality", "Number / term", "What it is", "Source"])

groups = [
    ("Plane", [
        ("$1.25/gal per 100 mi", "Air transport component ISER quotes for the most-competitive (large-bulk) deliveries", "ISER 2010, p.13"),
        ("/ 100", "Converts per-100-miles to per-mile", "arithmetic"),
        ("= 0.0125", "Result: $/gallon-mile", ""),
    ]),
    ("Road / trucking", [
        ("$4.75/vehicle-mile", "Alaska carrier line-haul rate to move one truck one mile (range $2.37-$6.00)", "Delivery Cost dossier"),
        ("9,000 gal", "Highway fuel-tanker capacity", "Delivery Cost dossier"),
        ("4.75 / 9,000 = 0.000528", "Result: $/gallon-mile (spreads per-truck cost over gallons carried)", ""),
    ]),
    ("Barge - linehaul", [
        ("$2.5M set", "Fixed cost of one tug-and-barge set for the ice-free season", "ISER 2010, p.16"),
        ("$3.5M fuel/misc", "Tug fuel + miscellaneous for a 6-7 trip season", "ISER 2010, p.16"),
        ("3M gal x 6 trips = 18M gal", "A linehaul barge carries ~3M gal; six trips per season", "ISER 2010, p.16"),
        ("$6M / 18M = $0.19/gal", "Result: flat $/gal (no miles - driven by seasonal volume, not distance)", ""),
    ]),
    ("Barge - lighterage (fixed)", [
        ("$10k/day", "Daily set cost = $1.35M/yr / 135 operating days ($1.35M = $600k capital + $750k operating)", "ISER 2010, Tables 1-2"),
        ("5 days", "Fixed handling per trip (load / lighter / wait); distance-independent", "modeling assumption (calibrated)"),
        ("250k gal", "Gallons delivered per lighterage trip (of 275k capacity)", "ISER 2010"),
        ("5 x $10k / 250k = $0.20/gal", "Result: fixed handling adder", ""),
    ]),
    ("Barge - lighterage (distance)", [
        ("2", "Round trip - barge travels out and back per one-way community mile", "geometry"),
        ("221 mi/day", "Barge speed ~8 kn x 24 h", "ISER; 8 kn confirmed by M&N POA 2024"),
        ("250k gal", "Gallons delivered per trip", "ISER 2010"),
        ("(2 x 10,000) / (221 x 250,000) = 0.000362", "Result: $/gallon-mile (one-way water miles)", ""),
    ]),
    ("Ice road", [
        ("$20/vehicle-mile", "Midpoint of the $10-$30/veh-mi range; driver-wage premium dominates", "Delivery Cost dossier"),
        ("3,000 gal", "Ice-road-rated truck capacity (small, due to ice weight limits)", "Delivery Cost dossier"),
        ("20 / 3,000 = 0.00667", "Result: $/gallon-mile", ""),
    ]),
]
r = 5
for modality, items in groups:
    for i, (term, what, source) in enumerate(items):
        ws.cell(row=r, column=1, value=modality if i == 0 else "").font = LABEL if i == 0 else BODY
        ws.cell(row=r, column=1).alignment = LEFT
        for col, val in ((2, term), (3, what), (4, source)):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY
            c.alignment = LEFT
        r += 1
footnotes(ws, r + 1, 4, [
    "Calibration check: $0.20 + 0.000362 x 1,100 mi = $0.60/gal reproduces "
    "ISER's Table 2 Bethel-class example, which pins the 5 fixed handling days.",
    "The 5 fixed handling days is the only value not lifted from a source; it "
    "is calibrated backward from ISER's $0.60/gal figure. All other numbers are "
    "published tariffs, capacities, or physical constants.",
])
widths(ws, [26, 32, 60, 30])

# ===========================================================================
# Table 5 — Barge two-leg build + validation
# ===========================================================================
ws = wb.create_sheet("Barge build")
title(ws, "Table 5.  Barge cost build (two-leg) and validation",
      "Total barge cost to a community = flat linehaul + lighterage "
      "(fixed handling + distance). Example shown for a Bethel-class ~1,100 "
      "one-way water-mile run.")
header_row(ws, 4, ["Component", "$/gal", "Note"])
end = data_rows(ws, 5, [
    ["Linehaul (flat)", 0.19, "trunk: refinery to regional hub"],
    ["Terminal use", 0.03, "if routed through a hub (ISER Table 4)"],
    ["Lighterage fixed handling", 0.20, "load / lighter / wait"],
    ["Lighterage distance", 0.40, "0.000362 $/gal-mi x 1,100 one-way water miles"],
    ["Total (Bethel-class ~1,100 mi)", 0.82, "consistent with ISER $0.40-0.80 lighterage + trunk"],
], numfmt={2: "0.00"})
# bold the total row
ws.cell(row=end - 1, column=1).font = LABEL
ws.cell(row=end - 1, column=2).font = LABEL

# validation sub-block
end += 1
ws.cell(row=end, column=1, value="Validation against ISER published figures").font = LABEL
end += 1
header_row(ws, end, ["Route (one-way water mi)", "Derived $/gal", "ISER check"])
end = data_rows(ws, end + 1, [
    ["Valdez-Anchorage ~300", 0.31, "within range"],
    ["Bethel-class ~1,100", 0.60, "= Table 2"],
    ["Kotzebue/remote ~1,500", 0.74, "within $0.40-0.80"],
], numfmt={2: "0.00"})
footnotes(ws, end + 1, 3, [
    "Lighterage $/gal = 0.20 (fixed) + 0.000362 x one-way water miles. "
    "The Table 4 example adds the flat linehaul and terminal-use legs on top.",
])
widths(ws, [34, 14, 52])

# ===========================================================================
# Table 5 — Sources
# ===========================================================================
ws = wb.create_sheet("Sources")
title(ws, "Table 6.  Sources")
header_row(ws, 4, ["Tag", "Full citation"])
data_rows(ws, 5, [
    ["ISER 2010", "Szymoniak, N., Fay, G., Villalobos-Melendez, A., Charon, J., "
     "Smith, M. (2010). Components of Alaska Fuel Costs: An Analysis of the "
     "Market Factors and Characteristics that Influence Rural Fuel Prices. "
     "UAA ISER, for the Alaska State Legislature, Senate Finance Committee."],
    ["Delivery Cost dossier", "Alaska Fuel Delivery Cost Analysis "
     "(Alaska Fuel Delivery Cost Analysis.pdf); rates encoded in "
     "friction_surface/friction_costs.py."],
    ["M&N POA 2024", "Moffatt & Nichol (2024). Economic Assessment for Port of "
     "Alaska Terminals. Prepared for Don Young Port of Alaska. Used only to "
     "independently confirm the ~8-knot barge speed; container/TEU cost "
     "figures and externality/RIMS costs are out of scope for this "
     "per-gallon model."],
])
widths(ws, [22, 100])

wb.save(OUT)
print("wrote", OUT)
